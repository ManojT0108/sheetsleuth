"""Parse an .xlsx workbook into a dependency graph: cells as nodes,
FEEDS_INTO edges from every formula reference (precedent -> dependent).

Output shape (JSON-serializable):
  nodes: [{id, sheet, address, row, col, kind, formula?, value?, literals?}]
  edges: [{src, dst}]                  # src FEEDS_INTO dst
kind is one of: formula | number | text | empty-ref (referenced but blank)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import get_column_letter, range_boundaries

MAX_RANGE_CELLS = 2000  # safety cap when expanding A1:Z999-style ranges


def _split_sheet_ref(ref: str, current_sheet: str) -> tuple[str, str]:
    """'Assumptions'!$B$3 -> ('Assumptions', '$B$3'); B3 -> (current, 'B3')."""
    if "!" in ref:
        sheet, addr = ref.rsplit("!", 1)
        return sheet.strip("'"), addr
    return current_sheet, ref


def _expand_ref(addr: str) -> list[str]:
    """Expand 'B3' or 'B3:B6' (possibly with $) into plain addresses."""
    min_col, min_row, max_col, max_row = range_boundaries(addr.replace("$", ""))
    if None in (min_col, min_row, max_col, max_row):
        return []  # whole-row/column ref; skip (not supported yet)
    n = (max_col - min_col + 1) * (max_row - min_row + 1)
    if n > MAX_RANGE_CELLS:
        return []
    return [
        f"{get_column_letter(c)}{r}"
        for c in range(min_col, max_col + 1)
        for r in range(min_row, max_row + 1)
    ]


def _parse_formula(formula: str, current_sheet: str):
    """Return (precedent cell ids, numeric literals) for a formula string."""
    precedents: list[str] = []
    literals: list[float] = []
    for tok in Tokenizer(formula).items:
        if tok.type != "OPERAND":
            continue
        if tok.subtype == "RANGE":
            sheet, addr = _split_sheet_ref(tok.value, current_sheet)
            precedents += [f"{sheet}!{a}" for a in _expand_ref(addr)]
        elif tok.subtype == "NUMBER":
            try:
                literals.append(float(tok.value))
            except ValueError:
                pass
    return precedents, literals


def extract_workbook(path: str | Path) -> dict:
    wb = load_workbook(path, data_only=False)
    nodes: dict[str, dict] = {}
    edges: list[dict] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cid = f"{ws.title}!{cell.coordinate}"
                node = {
                    "id": cid,
                    "sheet": ws.title,
                    "address": cell.coordinate,
                    "row": cell.row,
                    "col": cell.column,
                }
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    precedents, literals = _parse_formula(v, ws.title)
                    node["kind"] = "formula"
                    node["formula"] = v
                    if literals:
                        node["literals"] = literals
                    for p in precedents:
                        edges.append({"src": p, "dst": cid})
                elif isinstance(v, (int, float)):
                    node["kind"] = "number"
                    node["value"] = v
                else:
                    node["kind"] = "text"
                    node["value"] = str(v)
                nodes[cid] = node

    # referenced-but-blank cells still matter structurally
    for e in edges:
        if e["src"] not in nodes:
            sheet, addr = e["src"].rsplit("!", 1)
            nodes[e["src"]] = {
                "id": e["src"], "sheet": sheet, "address": addr,
                "kind": "empty-ref",
            }

    return {"nodes": list(nodes.values()), "edges": edges}


if __name__ == "__main__":
    import json
    import sys

    result = extract_workbook(sys.argv[1])
    kinds: dict[str, int] = {}
    for n in result["nodes"]:
        kinds[n["kind"]] = kinds.get(n["kind"], 0) + 1
    print(f"nodes: {len(result['nodes'])} {kinds}")
    print(f"edges: {len(result['edges'])}")
    out = Path(sys.argv[1]).with_suffix(".graph.json")
    out.write_text(json.dumps(result, indent=2))
    print(f"wrote {out}")
