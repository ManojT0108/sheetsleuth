"""Generate the SheetSleuth demo workbook: Lumeo Analytics FY2026 operating model.

A realistic seed-stage SaaS financial model with 5 planted errors, each a
classic real-world spreadsheet failure. Ground truth is emitted to
ground_truth.json so the verification engine can be proven against it.

Planted errors:
  E1 hardcoded-constant-in-chain: Revenue!G3 uses 0.15 growth typed during a
     what-if session instead of referencing Assumptions!B3 (0.08).
  E2 stale-pasted-constant: Revenue!J5 (Oct MRR) is a pasted number (September's
     value) replacing the formula — inconsistent with its row siblings.
  E3 short-sum-range: Costs!*8 totals use SUM(row3:row6), missing the
     contractors row 7 that was added later.
  E4 off-by-one-ref: Summary!B10 "Dec ARR (exit run-rate)" points at
     Revenue!L6 (November) instead of M6 (December).
  E5 orphaned-assumption: Assumptions!B12 (support cost per customer) feeds
     nothing — the founder believes support costs are modeled; they are not.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_COLS = [get_column_letter(c) for c in range(2, 14)]  # B..M

A = {  # assumptions: row -> (label, value, fmt)
    2: ("Starting customers (Jan)", 220, "0"),
    3: ("Monthly growth rate", 0.08, "0%"),
    4: ("Monthly churn rate", 0.03, "0%"),
    5: ("Price per seat / month ($)", 99, "$#,##0"),
    6: ("CAC — cost per new customer ($)", 1150, "$#,##0"),
    8: ("Monthly payroll ($)", 185000, "$#,##0"),
    9: ("Office & misc ($ / month)", 22000, "$#,##0"),
    10: ("Infra cost per customer ($ / month)", 11, "$#,##0"),
    11: ("Starting cash ($)", 2400000, "$#,##0"),
    12: ("Support cost per customer ($ / month)", 6, "$#,##0"),  # E5: orphaned
}

CONTRACTORS = {"Mar": 15000, "Jul": 8000, "Oct": 12000}

HEADER_FILL = PatternFill("solid", fgColor="1F3557")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def simulate():
    """Recreate the workbook's math in Python (including E1) so pasted
    constants (E2) are internally consistent with the broken model."""
    growth, churn = A[3][1], A[4][1]
    customers = [A[2][1]]
    for i in range(1, 12):
        g = 0.15 if MONTHS[i] == "Jun" else growth  # E1 lives in June
        customers.append(round(customers[i - 1] * (1 + g - churn)))
    mrr = [c * A[5][1] for c in customers]
    return customers, mrr


def style_header(ws, row, last_col=13):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def build():
    customers, mrr = simulate()
    sept_mrr = mrr[MONTHS.index("Sep")]  # E2: the stale paste

    wb = Workbook()

    # ---------------- Assumptions ----------------
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = "Lumeo Analytics — FY2026 Operating Model Assumptions"
    ws["A1"].font = Font(bold=True, size=14)
    for row, (label, value, fmt) in A.items():
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = fmt
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14

    # ---------------- Revenue ----------------
    ws = wb.create_sheet("Revenue")
    ws["A1"] = "Revenue Build"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Month"
    for i, col in enumerate(MONTH_COLS):
        ws[f"{col}2"] = MONTHS[i]
    style_header(ws, 2)

    ws["A3"] = "Customers (EOM)"
    ws["B3"] = "=Assumptions!B2"
    for i in range(1, 12):
        col, prev = MONTH_COLS[i], MONTH_COLS[i - 1]
        if MONTHS[i] == "Jun":  # E1: hardcoded 0.15 growth
            ws[f"{col}3"] = f"=ROUND({prev}3*(1+0.15-Assumptions!$B$4),0)"
        else:
            ws[f"{col}3"] = (
                f"=ROUND({prev}3*(1+Assumptions!$B$3-Assumptions!$B$4),0)"
            )

    ws["A4"] = "New customers"
    ws["B4"] = "=ROUND(Assumptions!B2*Assumptions!$B$3,0)"
    for i in range(1, 12):
        col, prev = MONTH_COLS[i], MONTH_COLS[i - 1]
        ws[f"{col}4"] = f"=ROUND({prev}3*Assumptions!$B$3,0)"

    ws["A5"] = "MRR ($)"
    for i, col in enumerate(MONTH_COLS):
        if MONTHS[i] == "Oct":  # E2: pasted September value
            ws[f"{col}5"] = sept_mrr
        else:
            ws[f"{col}5"] = f"={col}3*Assumptions!$B$5"

    ws["A6"] = "ARR run-rate ($)"
    for col in MONTH_COLS:
        ws[f"{col}6"] = f"={col}5*12"

    for row, fmt in ((3, "0"), (4, "0"), (5, "$#,##0"), (6, "$#,##0")):
        for col in MONTH_COLS:
            ws[f"{col}{row}"].number_format = fmt
    ws.column_dimensions["A"].width = 20

    # ---------------- Costs ----------------
    ws = wb.create_sheet("Costs")
    ws["A1"] = "Cost Build"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Month"
    for i, col in enumerate(MONTH_COLS):
        ws[f"{col}2"] = MONTHS[i]
    style_header(ws, 2)

    rows = {
        3: ("Payroll ($)", "=Assumptions!$B$8"),
        4: ("Marketing / CAC ($)", "=Revenue!{c}4*Assumptions!$B$6"),
        5: ("Infrastructure ($)", "=Revenue!{c}3*Assumptions!$B$10"),
        6: ("Office & misc ($)", "=Assumptions!$B$9"),
    }
    for r, (label, tmpl) in rows.items():
        ws[f"A{r}"] = label
        for col in MONTH_COLS:
            ws[f"{col}{r}"] = tmpl.format(c=col)

    ws["A7"] = "Contractors & one-time ($)"  # added later; E3 misses it
    for i, col in enumerate(MONTH_COLS):
        ws[f"{col}7"] = CONTRACTORS.get(MONTHS[i], 0)

    ws["A8"] = "Total costs ($)"
    for col in MONTH_COLS:
        ws[f"{col}8"] = f"=SUM({col}3:{col}6)"  # E3: short range

    for r in range(3, 9):
        for col in MONTH_COLS:
            ws[f"{col}{r}"].number_format = "$#,##0"
    ws.column_dimensions["A"].width = 24

    # ---------------- Summary ----------------
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Board Summary — FY2026"
    ws["A1"].font = Font(bold=True, size=14)
    items = {
        3: ("Total FY2026 revenue ($)", "=SUM(Revenue!B5:M5)", "$#,##0"),
        4: ("Total FY2026 costs ($)", "=SUM(Costs!B8:M8)", "$#,##0"),
        5: ("Net income ($)", "=B3-B4", "$#,##0"),
        6: ("Ending cash ($)", "=Assumptions!B11+B5", "$#,##0"),
        7: ("Avg monthly burn ($)", "=(B4-B3)/12", "$#,##0"),
        8: ("Runway (months)", '=IF(B5>0,"Profitable",Assumptions!B11/((B4-B3)/12))', "0.0"),
        10: ("Dec ARR — exit run-rate ($)", "=Revenue!L6", "$#,##0"),  # E4: L6 not M6
        11: ("Customers at exit", "=Revenue!M3", "0"),
    }
    for r, (label, formula, fmt) in items.items():
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=formula)
        c.number_format = fmt
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16

    out = OUT_DIR / "lumeo_fy2026_model.xlsx"
    wb.save(out)
    return out, sept_mrr


GROUND_TRUTH = [
    {
        "id": "E1",
        "type": "hardcoded-constant-in-chain",
        "sheet": "Revenue", "cells": ["G3"],
        "story": "A 15% growth rate typed into June's customer formula during a "
                 "what-if session and never reverted; every month after June is "
                 "inflated, propagating to MRR, ARR, marketing costs, and the "
                 "board summary.",
        "expected_fix": "Replace literal 0.15 with Assumptions!$B$3 (8%).",
    },
    {
        "id": "E2",
        "type": "stale-pasted-constant",
        "sheet": "Revenue", "cells": ["K5"],
        "story": "October MRR was overwritten with a pasted number (September's "
                 "value) during board-deck prep; the formula never came back.",
        "expected_fix": "Restore =J3*Assumptions!$B$5.",
    },
    {
        "id": "E3",
        "type": "short-sum-range",
        "sheet": "Costs", "cells": [f"{c}8" for c in MONTH_COLS],
        "story": "The contractors row was added after the totals row; SUM still "
                 "ends at row 6, so $35,000 of spend is invisible to the total.",
        "expected_fix": "Extend SUM(B3:B6) to SUM(B3:B7).",
    },
    {
        "id": "E4",
        "type": "off-by-one-reference",
        "sheet": "Summary", "cells": ["B10"],
        "story": "'Dec ARR' points at Revenue!L6 — November — a classic "
                 "drag-fill off-by-one. The board sees the wrong exit run-rate.",
        "expected_fix": "Point B10 at Revenue!M6.",
    },
    {
        "id": "E5",
        "type": "orphaned-assumption",
        "sheet": "Assumptions", "cells": ["B12"],
        "story": "Support cost per customer exists as an assumption but feeds "
                 "nothing; the founder believes support costs are modeled. "
                 "They are not — costs are structurally understated.",
        "expected_fix": "Wire B12 into the Costs sheet (customers x support cost).",
    },
]

if __name__ == "__main__":
    path, sept_mrr = build()
    gt = OUT_DIR / "ground_truth.json"
    gt.write_text(json.dumps(GROUND_TRUTH, indent=2))
    print(f"wrote {path}")
    print(f"wrote {gt}")
    print(f"E2 pasted value (Sep MRR): ${sept_mrr:,}")
