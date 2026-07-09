"""Integration: loader + the 5 smell detectors against a real Neo4j.
Skipped when no Neo4j is reachable. Uses a throwaway workbook id and
cleans up after itself."""

import uuid

import pytest

from tests.conftest import needs_neo4j

WB = f"pytest-{uuid.uuid4().hex[:8]}"


@pytest.fixture(scope="module")
def loaded(demo_xlsx):
    from app.graph.loader import clear_workbook, load_workbook
    from app.parser.extract import extract_workbook
    stats = load_workbook(WB, "pytest demo", extract_workbook(demo_xlsx))
    yield stats
    clear_workbook(WB)


@needs_neo4j
def test_load_stats(loaded):
    assert loaded == {"cells": 196, "edges": 247, "sheets": 4}


@needs_neo4j
def test_detectors_find_exact_ground_truth(loaded):
    from app.audit import smells
    findings = smells.detect(WB)
    by_type = {f["type"]: f for f in findings}
    assert set(by_type) == {
        "hardcoded-constant-in-chain", "stale-pasted-constant",
        "short-sum-range", "orphaned-assumption",
    }  # E4 is the LLM's job; no circular refs in the demo model

    e1 = by_type["hardcoded-constant-in-chain"]
    assert e1["cells"] == [f"{WB}::Revenue!G3"]
    assert e1["evidence"]["suspects"] == [0.15]
    assert e1["evidence"]["blast"] == 52

    e2 = by_type["stale-pasted-constant"]
    assert e2["cells"] == [f"{WB}::Revenue!K5"]

    e3 = by_type["short-sum-range"]
    assert len(e3["cells"]) == 12
    assert e3["evidence"]["missedStaticValue"] == 35000

    e5 = by_type["orphaned-assumption"]
    assert e5["cells"] == [f"{WB}::Assumptions!B12"]


@needs_neo4j
def test_blast_radius_and_critical_cells(loaded):
    from app.graph import queries as q
    br = q.blast_radius(f"{WB}::Revenue!G3")
    assert br["count"] == 52
    assert set(br["sheetsReached"]) == {"Costs", "Revenue", "Summary"}

    crit = q.critical_cells(WB, 3)
    assert crit[0]["cell"] == f"{WB}::Assumptions!B2"  # starting customers


@needs_neo4j
def test_circular_reference_detector():
    """Load a tiny synthetic workbook with A1<->B1 and expect detection."""
    import openpyxl

    from app.audit import smells
    from app.graph.loader import clear_workbook, load_workbook
    from app.parser.extract import extract_workbook
    import tempfile
    from pathlib import Path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=B1+1"
    ws["B1"] = "=A1+1"
    path = Path(tempfile.mkdtemp()) / "circular.xlsx"
    wb.save(path)

    wbid = f"pytest-circ-{uuid.uuid4().hex[:6]}"
    try:
        load_workbook(wbid, "circular", extract_workbook(path))
        findings = smells.detect(wbid)
        assert any(f["type"] == "circular-reference" for f in findings)
    finally:
        clear_workbook(wbid)
